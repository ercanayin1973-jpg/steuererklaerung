import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# İsviçre/Zürih 2025 Kurallarına Göre Profesyonel Veri Yapısı
data = {
    "Ziffer": ["11.1", "11.2", "17", "21", "22.1", "22.2", "30.1", "35", "16.6"],
    "Beschreibung": [
        "Bruttolohn P1 (Lohnausweis)", "Bruttolohn P2 (Lohnausweis)", 
        "Sonderabzug Erwerbstätigkeit (Beide)", "Berufsauslagen (P1+P2)", 
        "Säule 3a (Max 7'258 x 2)", "Krankenkassen-Sozialabzug", 
        "Bankkonten (ZKB)", "Schuldenverzeichnis (Privatkredit)", "Kita-Abzug (Fremdbetreuung)"
    ],
    "Betrag (CHF)": [125000, 45000, 6100, -6400, -14516, -5800, 45000, -5000, -12000]
}

df = pd.DataFrame(data)

# Excel Dosyası Oluştur
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Steuererklaerung_Zuerich_2025"

# Header ve Styling
header = ["Ziffer", "Beschreibung", "Betrag (CHF)"]
for col_num, title in enumerate(header, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = title
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    ws.column_dimensions[get_column_letter(col_num)].width = 30

# Verileri Yaz
for r_idx, row in enumerate(df.values, 2):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx).value = value

# Hesaplama - Steuerbares Einkommen
total_inc = df[df['Ziffer'].isin(['11.1', '11.2'])]['Betrag (CHF)'].sum()
total_ded = df[df['Ziffer'].isin(['21', '22.1', '22.2', '16.6'])]['Betrag (CHF)'].sum()
steuerbar = total_inc + total_ded

ws.cell(row=10, column=2).value = "Steuerbares Einkommen (Rechnung):"
ws.cell(row=10, column=3).value = steuerbar
ws.cell(row=10, column=3).font = Font(bold=True, color="B91C1C")

wb.save("Steuererklaerung_Zuerich_2025_Detail.xlsx")
